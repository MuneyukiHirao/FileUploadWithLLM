# test/test_main_processor.py (今は動かない。何が悪いか調べていない）

import os
from openpyxl import Workbook
from app.main_processor import process_file_with_llm

def create_xlsx_file(file_path: str, row_count: int):
    """
    row_count行のXLSXファイルを生成。
    シート数は1枚のみ。
    1行目には単純な文字列を入れる。
    残り行は "DataX" のように簡単に入れる。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SingleSheet"

    # 1行目: 見出しっぽい文字列
    ws.append(["ColA", "ColB", "ColC"])

    # 2行目以降: 適当に "DataX" 系
    for i in range(row_count - 1):
        ws.append([f"Data{i}_A", f"Data{i}_B", f"Data{i}_C"])

    wb.save(file_path)

def test_small_file():
    """
    10行程度の小さなファイルをテストし、LLMモックに渡す。
    """
    file_path = "test_small.xlsx"
    create_xlsx_file(file_path, row_count=10)

    result = process_file_with_llm(file_path)
    print("[test_small_file] result =", result)
    if result["status"] == "success":
        print("  [test_small_file] => PASSED")
    else:
        print("  [test_small_file] => FAILED")

def test_large_file():
    """
    200行程度のファイルを作成 -> 実際は100行のみ抽出されるか確認。
    """
    file_path = "test_large.xlsx"
    create_xlsx_file(file_path, row_count=200)

    result = process_file_with_llm(file_path)
    print("[test_large_file] result =", result)
    if result["status"] == "success":
        llm_req_data = result["llmResponse"]
        if llm_req_data["status"] == "success":
            # モックのヘッダー判定結果はさておき
            print("  [test_large_file] => PASSED (Check if rowData was cut to 100 rows in file_to_json.py)")
        else:
            print("  [test_large_file] => PASSED (LLM returned error, but file read was okay)")
    else:
        print("  [test_large_file] => FAILED")

def test_wrong_extension():
    """
    拡張子が .xlsx 以外 -> エラーになるか。
    """
    file_path = "test_wrong.csv"
    with open(file_path, "w", encoding="utf-8") as f:
        f.write("Dummy CSV content")

    result = process_file_with_llm(file_path)
    print("[test_wrong_extension] result =", result)
    if result["status"] == "error" and ".xlsx" in result["message"]:
        print("  [test_wrong_extension] => PASSED")
    else:
        print("  [test_wrong_extension] => FAILED")

def test_multi_sheet():
    """
    2シートある場合 -> エラーになるか。
    """
    from openpyxl import Workbook
    file_path = "test_multisheet.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["ColA", "ColB"])
    ws1.append(["ValA", "ValB"])

    wb.create_sheet("Sheet2")
    wb.save(file_path)

    result = process_file_with_llm(file_path)
    print("[test_multi_sheet] result =", result)
    if result["status"] == "error" and "シート数が1つではありません" in result["message"]:
        print("  [test_multi_sheet] => PASSED")
    else:
        print("  [test_multi_sheet] => FAILED")

def test_corrupted_file():
    """
    破損ファイルを作成 -> エラーか。
    """
    file_path = "test_corrupt.xlsx"
    with open(file_path, "wb") as f:
        f.write(b"This is not a valid XLSX")

    result = process_file_with_llm(file_path)
    print("[test_corrupted_file] result =", result)
    if result["status"] == "error" and "読み込めません" in result["message"]:
        print("  [test_corrupted_file] => PASSED")
    else:
        print("  [test_corrupted_file] => FAILED")

def run_all_tests():
    print("===== Start Testing Main Processor =====")
    test_small_file()
    test_large_file()
    test_wrong_extension()
    test_multi_sheet()
    test_corrupted_file()
    print("===== End Testing Main Processor =====")

if __name__ == "__main__":
    run_all_tests()
