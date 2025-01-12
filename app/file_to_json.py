# app/file_to_json.py

import os
from openpyxl import load_workbook
from typing import Dict, Any

MAX_FILE_SIZE = 500 * 1024 * 1024  # 500MB
MAX_ROWS = 100

def convert_xlsx_to_json(file_path: str) -> Dict[str, Any]:
    """
    大きなXLSXファイルが来ても、先頭MAX_ROWS行のみを読み込み、
    それをrowDataとしてJSON形式のdictにまとめて返す。
    
    ヘッダー判定は実施せず、生データを返すことに特化。
    """

    if not file_path.lower().endswith('.xlsx'):
        return {
            "status": "error",
            "message": f"拡張子が .xlsx ではありません: {file_path}"
        }

    if os.path.getsize(file_path) > MAX_FILE_SIZE:
        return {
            "status": "error",
            "message": "ファイルサイズが500MBを超えています。"
        }

    try:
        wb = load_workbook(filename=file_path, read_only=True)
    except Exception as e:
        return {
            "status": "error",
            "message": f"ファイルが破損しているか、読み込めません: {str(e)}"
        }

    sheetnames = wb.sheetnames
    if len(sheetnames) != 1:
        return {
            "status": "error",
            "message": f"シート数が1つではありません。シート数: {len(sheetnames)}"
        }

    ws = wb[sheetnames[0]]

    rowData = []
    count = 0
    for row in ws.iter_rows(values_only=True):
        rowData.append(list(row))
        count += 1
        if count >= MAX_ROWS:
            break

    return {
        "status": "success",
        "fileType": "xlsx",
        "rowData": rowData
    }